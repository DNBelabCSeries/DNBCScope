#!/usr/bin/env python3
"""DNBCScope cell annotation backends: ScType-like markers and lightweight CellTypist inference."""

import json
import os
import pickle
import struct
import sys
import warnings
from functools import lru_cache
from typing import Dict, List, Tuple


def emit_json_result(value):
    """Use the result path supplied by the resident worker when available."""
    result_path = os.environ.get("DNBC_RESULT_PATH")
    if result_path:
        # Compact annotation results use a binary frame so a million label IDs
        # and float32 scores are not expanded into a second JSON string. Keep
        # the JSON writer for old workers/fixtures and for direct CLI output.
        if "cell_label_ids" in value and "label_dictionary" in value:
            _write_compact_result_frame(result_path, value)
            return
        encoded = json.dumps(value, separators=(",", ":"), allow_nan=False)
        with open(result_path, "w", encoding="utf-8") as handle:
            handle.write(encoded)
    else:
        encoded = json.dumps(value, separators=(",", ":"), allow_nan=False)
        print(encoded)


# Redirect all warnings to stderr to keep stdout clean for JSON output
warnings.filterwarnings("default")
warnings.showwarning = lambda msg, cat, fn, ln, file=None, line=None: sys.stderr.write(
    f"{fn}:{ln}: {cat.__name__}: {msg}\n"
)

# Keep direct execution and the resident scientific worker on the same import
# contract. Embedded Windows Python may not add the script directory when its
# ``python._pth`` file enables isolated mode.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

try:
    from expression_source import make_expression_adata, normalize_windows_path, read_10x_mtx_flexible
except ImportError as error:
    raise RuntimeError(
        "DNBCScope could not load expression_source.py next to run_annotation.py"
    ) from error

def select_expression_adata(adata, cell_indices=None):
    selected, is_counts, _source_label = make_expression_adata(
        adata, allow_negative=True, cell_indices=cell_indices
    )
    return selected, is_counts


def resolve_h5ad(path: str) -> str:
    path = normalize_windows_path(path)
    if os.path.isfile(path) and path.lower().endswith(".h5ad"):
        return path
    # Match the analysis/expression loaders: source choice must not depend on
    # filesystem enumeration order (which differs on macOS and Windows).
    for name in sorted(os.listdir(path), key=lambda value: (value.casefold(), value)):
        if name.lower().endswith(".h5ad"):
            return os.path.join(path, name)
    raise RuntimeError(f"No h5ad file found in {path}")


def load_source(path: str, data_format: str):
    path = normalize_windows_path(path)
    import anndata as ad

    if data_format == "native-csr":
        from native_csr import load_native_csr
        return load_native_csr(path, include_barcodes=False)
    if data_format == "h5ad":
        return ad.read_h5ad(resolve_h5ad(path), backed="r")
    if data_format != "mtx":
        raise RuntimeError(f"Unsupported data format: {data_format}")
    # Keep annotation on the same tolerant MTX loader as analysis and
    # differential expression.  The previous local reader required only a
    # feature file, did not validate barcodes/dimensions, and left duplicate
    # symbols to downstream model matching; a project could therefore import
    # successfully but fail or annotate different genes in this step.
    return read_10x_mtx_flexible(path)


def normalize(adata, normalize_input=True):
    import scanpy as sc

    adata = adata.copy()
    adata.var_names_make_unique()
    if normalize_input:
        sc.pp.normalize_total(adata, target_sum=1e4)
        sc.pp.log1p(adata)
    return adata


def to_dense(matrix):
    if hasattr(matrix, "toarray"):
        return matrix.toarray()
    return matrix


def _validated_cluster_ids(values, n_cells, cluster_count):
    """Validate JSON/sidecar cluster ids before any narrowing cast."""
    import numpy as np

    raw = np.asarray(values)
    if raw.ndim != 1 or raw.size != n_cells:
        raise RuntimeError("Cluster IDs do not match the number of analyzed cells")
    if raw.size:
        if raw.dtype.kind == "b":
            raise RuntimeError("Cluster IDs must be integers")
        if raw.dtype.kind not in "iu":
            if raw.dtype.kind not in "fc" or not np.isfinite(raw).all() or not np.equal(raw, np.floor(raw)).all():
                raise RuntimeError("Cluster IDs must be finite integers")
        if np.any(raw < 0) or np.any(raw >= cluster_count):
            raise RuntimeError("Cluster IDs contain an out-of-range value")
    return raw.astype(np.intp, copy=False)


def _sample_expression_rows(matrix, max_rows=64):
    """Read a bounded, dense row sample without materializing a full h5ad.

    CellTypist's AnnData contract is expressed in terms of per-cell values
    (log1p-normalized to a target library size), so validating the contract
    should not require loading every cell of a large project.
    """
    import numpy as np
    import scipy.sparse as sp

    if not hasattr(matrix, "shape") or len(matrix.shape) != 2:
        return np.asarray(matrix)
    n_rows = int(matrix.shape[0])
    if n_rows == 0:
        return np.empty((0, int(matrix.shape[1])), dtype=np.float64)
    row_indices = np.linspace(
        0,
        n_rows - 1,
        num=min(max_rows, n_rows),
        dtype=np.int64,
    )
    row_indices = np.unique(row_indices)
    sampled = matrix[row_indices]
    if hasattr(sampled, "to_memory"):
        sampled = sampled.to_memory()
    if sp.issparse(sampled):
        sampled = sampled.toarray()
    return np.asarray(sampled)


def validate_celltypist_expression(adata, target_sum=1e4):
    """Validate the expression scale expected by official CellTypist.

    CellTypist does not normalize an AnnData input.  It expects non-negative
    log1p values whose inverse-transformed library size is approximately
    10,000; raw counts must be normalized/log-transformed before this check.
    The upstream implementation warns when the library size differs, but it
    rejects negative or clearly un-normalized values.  We follow that contract
    while making the warning explicit in the worker stderr.
    """
    import numpy as np

    sampled = _sample_expression_rows(adata.X)
    if sampled.size == 0:
        return
    if not np.isfinite(sampled).all() or float(np.min(sampled)) < 0:
        raise RuntimeError(
            "CellTypist requires non-negative log1p-normalized expression. "
            "The selected matrix contains invalid or negative values."
        )

    # The official AnnData path treats values above log1p(10000) as an invalid
    # expression source (and would fall back to .raw when available).  At this
    # point expression_source has already selected the best available source,
    # so fail with an actionable message rather than silently scoring counts or
    # linear CPM values as if they were log1p values.
    if float(np.max(sampled)) > 9.22:
        raise RuntimeError(
            "CellTypist requires log1p-normalized expression (target sum 10,000). "
            "The selected matrix contains values above the supported range; "
            "provide raw counts in X, layers['counts'], or raw.X."
        )

    inverse_library_sizes = np.expm1(sampled).sum(axis=1)
    if not np.allclose(inverse_library_sizes, target_sum, atol=1.0, rtol=0.0):
        warnings.warn(
            "CellTypist expects all genes in log1p-normalized expression with "
            f"an inverse-transformed library size of {target_sum:g}; the input "
            "does not satisfy this for every sampled cell. Predictions may be "
            "less reliable when genes were filtered or another normalization "
            "scale was used.",
            RuntimeWarning,
        )


def _match_celltypist_features(data_names, model_features):
    """Return data/model column indices using CellTypist's exact gene match.

    The upstream classifier uses ``np.isin`` with exact feature strings and
    preserves the input-gene order.  Case-insensitive matching is tempting for
    convenience, but it can silently match human ``CD3D`` to mouse ``Cd3d``
    and produce a confident result from the wrong species model.
    """
    model_lookup = {}
    for index, feature in enumerate(model_features):
        key = str(feature)
        if key in model_lookup:
            raise RuntimeError(f"CellTypist model contains duplicate feature gene '{key}'")
        model_lookup[key] = index

    model_indices = []
    data_indices = []
    for data_index, name in enumerate(data_names):
        model_index = model_lookup.get(str(name))
        if model_index is not None:
            data_indices.append(data_index)
            model_indices.append(model_index)
    return model_indices, data_indices


def load_sctype_db():
    db_path = os.environ.get("DNBCSCOPE_SCTYPE_DB_PATH")
    if not db_path or not os.path.exists(db_path):
        raise RuntimeError("ScType DB is not available")
    return _load_sctype_db_cached(os.path.abspath(normalize_windows_path(db_path)))


@lru_cache(maxsize=2)
def _load_sctype_db_cached(db_path: str):
    """Cache by resolved path so a resident worker can switch databases."""
    ext = os.path.splitext(db_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _load_sctype_db_from_excel(db_path)
    with open(db_path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def _load_sctype_db_from_excel(path: str) -> dict:
    """Load ScType DB from Excel file (ScTypeDB_full.xlsx format)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required to read Excel marker files. Install it with: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    header = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    columns = {name.lower(): index for index, name in enumerate(header)}
    tissue_col = columns.get("tissuetype", columns.get("tissue", 0))
    cell_col = columns.get("cellname", columns.get("cell_type", 1))
    positive_col = columns.get("genesymbolmore1", columns.get("markers", 2))
    negative_col = columns.get("genesymbolmore2", columns.get("negative", 3))
    short_col = columns.get("shortname", columns.get("short_name", 4))
    tissues = set()
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if tissue_col >= len(row) or not row[tissue_col]:
            continue
        tissue = str(row[tissue_col]).strip()
        cell_type = str(row[cell_col]).strip() if cell_col < len(row) and row[cell_col] else ""
        markers = str(row[positive_col]).strip() if positive_col < len(row) and row[positive_col] else ""
        negative = str(row[negative_col]).strip() if negative_col < len(row) and row[negative_col] else ""
        short_name = str(row[short_col]).strip() if short_col < len(row) and row[short_col] else cell_type
        if not cell_type and not short_name:
            continue
        tissues.add(tissue)
        pos_genes = [g.strip() for g in markers.split(",") if g.strip()] if markers else []
        neg_genes = [g.strip() for g in negative.split(",") if g.strip()] if negative else []
        entries.append({
            "tissue": tissue,
            "label": cell_type or short_name,
            "cell_name": cell_type,
            "short_name": short_name,
            "positive": pos_genes,
            "negative": neg_genes,
        })
    wb.close()
    return {"tissues": sorted(tissues), "entries": entries}


def get_marker_db(tissue: str) -> Dict[str, Dict[str, List[str]]]:
    db = load_sctype_db()
    normalized_tissue = (tissue or "").strip().lower()
    result: Dict[str, Dict[str, List[str]]] = {}
    for entry in db.get("entries", []):
        entry_tissue = str(entry.get("tissue", "")).strip().lower()
        if normalized_tissue and entry_tissue != normalized_tissue:
            continue
        label = str(entry.get("label") or entry.get("short_name") or "").strip()
        if not label:
            continue
        # Merge duplicate labels within the same tissue
        if label in result:
            result[label]["positive"].extend(entry.get("positive", []))
            result[label]["negative"].extend(entry.get("negative", []))
        else:
            result[label] = {
                "positive": list(entry.get("positive", [])),
                "negative": list(entry.get("negative", [])),
            }
    if not result:
        available = ", ".join(load_sctype_db().get("tissues", [])[:12])
        raise RuntimeError(f"No ScType markers found for tissue '{tissue}'. Available tissues include: {available}")
    return result


def _zscore_rows(matrix):
    import numpy as np

    means = matrix.mean(axis=1, keepdims=True)
    stds = matrix.std(axis=1, keepdims=True)
    stds[stds == 0] = 1.0
    return (matrix - means) / stds


def _compute_marker_sensitivity(marker_db: Dict[str, Dict[str, List[str]]]) -> Dict[str, float]:
    """Compute marker sensitivity as in original scType.

    Score = 1 - (count - 1) / (N_celltypes - 1)
    A gene appearing in only 1 cell type -> score = 1 (most specific)
    A gene appearing in all cell types -> score = 0 (least specific)
    """
    marker_counts: Dict[str, int] = {}
    for entry in marker_db.values():
        for gene in set(list(entry.get("positive", [])) + list(entry.get("negative", []))):
            gene_upper = gene.strip().upper()
            if gene_upper:
                marker_counts[gene_upper] = marker_counts.get(gene_upper, 0) + 1
    n_celltypes = len(marker_db)
    if n_celltypes <= 1:
        return {gene: 1.0 for gene in marker_counts}
    return {
        gene: 1.0 - (count - 1) / (n_celltypes - 1)
        for gene, count in marker_counts.items()
    }


def _dense_expression_chunk(adata, start: int, stop: int, marker_indices):
    """Load one cells-by-markers block, never the whole marker matrix."""
    import numpy as np
    import scipy.sparse as sp

    block = adata[start:stop, marker_indices].X
    if hasattr(block, "to_memory"):
        block = block.to_memory()
    if sp.issparse(block):
        block = block.toarray()
    block = np.asarray(block, dtype=np.float32)
    if block.ndim != 2:
        raise RuntimeError("Expected a 2D expression matrix for ScType annotation")
    return block


def run_sctype(adata, tissue: str, cluster_ids=None, cluster_names=None, chunk_size: int = 8192):
    """Run scType without retaining a cells-by-label score matrix.

    The old implementation kept both ``cells x markers`` and ``cells x labels``
    arrays alive.  On a million-cell project that can consume multiple GB.  We
    now make two bounded passes over the marker columns: the first computes the
    per-marker z-score parameters, and the second scores one cell block at a
    time.  Cluster consensus only needs ``clusters x labels`` sums, which is
    small and is retained for the existing cluster-level semantics.
    """
    import math
    import numpy as np

    marker_db = get_marker_db(tissue)
    if not marker_db:
        raise RuntimeError(f"No ScType marker database available for tissue {tissue}")

    var_lookup = {str(name).upper(): idx for idx, name in enumerate(adata.var_names)}
    marker_genes = sorted({
        gene.strip().upper()
        for entry in marker_db.values()
        for gene in list(entry.get("positive", [])) + list(entry.get("negative", []))
        if gene.strip().upper() in var_lookup
    })
    if not marker_genes:
        raise RuntimeError(f"No marker genes from the ScType tissue '{tissue}' set were found in this dataset")

    marker_indices = [var_lookup[gene] for gene in marker_genes]
    marker_lookup = {gene: idx for idx, gene in enumerate(marker_genes)}
    n_cells = int(adata.n_obs)
    if n_cells <= 0:
        raise RuntimeError("ScType requires at least one cell")

    # Compute marker sensitivity (original: score = 1 - (count-1)/(N-1)).
    marker_sensitivity = _compute_marker_sensitivity(marker_db)
    sensitivity = np.asarray(
        [marker_sensitivity.get(gene, 1.0) for gene in marker_genes], dtype=np.float32
    )

    # Collect marker indices once.  Keeping indices rather than expression
    # slices lets every chunk be released before the next one is read.
    all_pos: Dict[str, List[int]] = {}
    all_neg: Dict[str, List[int]] = {}
    used_markers: set = set()
    for label, entry in marker_db.items():
        pos_found = [g.upper() for g in entry.get("positive", []) if g.upper() in marker_lookup]
        neg_found = [g.upper() for g in entry.get("negative", []) if g.upper() in marker_lookup]
        all_pos[label] = [marker_lookup[g] for g in pos_found]
        all_neg[label] = [marker_lookup[g] for g in neg_found]
        used_markers.update(pos_found)
        used_markers.update(neg_found)

    labels = list(all_pos.keys())
    n_labels = len(labels)
    if n_labels == 0:
        raise RuntimeError("No ScType labels are available for the selected tissue")

    # First bounded pass: z-score parameters per marker.  Float64 accumulators
    # avoid cancellation for large cell counts; the scored blocks stay float32.
    sums = np.zeros(len(marker_indices), dtype=np.float64)
    sums_sq = np.zeros(len(marker_indices), dtype=np.float64)
    for start in range(0, n_cells, max(1, int(chunk_size))):
        stop = min(start + max(1, int(chunk_size)), n_cells)
        block = _dense_expression_chunk(adata, start, stop, marker_indices)
        block64 = block.astype(np.float64, copy=False)
        sums += block64.sum(axis=0)
        sums_sq += np.square(block64).sum(axis=0)
    means64 = sums / n_cells
    means = means64.astype(np.float32)
    variance = np.maximum(sums_sq / n_cells - np.square(means64), 0.0)
    stds = np.sqrt(variance).astype(np.float32)
    stds[stds == 0] = 1.0

    cluster_array = None
    cluster_score_sums = None
    cluster_cell_counts = None
    if cluster_ids is not None and cluster_names:
        cluster_count = len(cluster_names)
        cluster_array = _validated_cluster_ids(cluster_ids, n_cells, cluster_count)
        cluster_score_sums = np.zeros((cluster_count, n_labels), dtype=np.float64)
        cluster_cell_counts = np.bincount(cluster_array, minlength=cluster_count)

    cell_label_ids = np.empty(n_cells, dtype=np.uint32)
    best_scores = np.empty(n_cells, dtype=np.float32)
    confidence_gaps = np.empty(n_cells, dtype=np.float32)
    for start in range(0, n_cells, max(1, int(chunk_size))):
        stop = min(start + max(1, int(chunk_size)), n_cells)
        block = _dense_expression_chunk(adata, start, stop, marker_indices)
        z = (block - means) / stds
        z *= sensitivity
        score_block = np.zeros((stop - start, n_labels), dtype=np.float32)
        for label_index, label in enumerate(labels):
            pos_idx = all_pos[label]
            neg_idx = all_neg[label]
            if pos_idx:
                score_block[:, label_index] += z[:, pos_idx].sum(axis=1) / math.sqrt(len(pos_idx))
            if neg_idx:
                score_block[:, label_index] -= z[:, neg_idx].sum(axis=1) / math.sqrt(len(neg_idx))

        best_idx = score_block.argmax(axis=1)
        best_score = score_block[np.arange(stop - start), best_idx]
        if n_labels > 1:
            second_score = np.partition(score_block, -2, axis=1)[:, -2]
        else:
            second_score = np.zeros(stop - start, dtype=np.float32)
        cell_label_ids[start:stop] = best_idx.astype(np.uint32, copy=False)
        best_scores[start:stop] = best_score
        confidence_gaps[start:stop] = best_score - second_score

        if cluster_score_sums is not None:
            block_clusters = cluster_array[start:stop]
            np.add.at(cluster_score_sums, block_clusters, score_block)

    return {
        "cell_label_ids": cell_label_ids,
        "cell_scores_f32": best_scores,
        "cell_confidences_f32": confidence_gaps,
        "label_dictionary": labels,
        "evidence_count": len(used_markers),
        "score_labels": labels,
        "cluster_score_sums": cluster_score_sums,
        "cluster_cell_counts": cluster_cell_counts,
    }


class _CellTypistModelStub:
    """Stub class to unpickle CellTypist models without importing celltypist."""
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)

    def __setstate__(self, state):
        for k, v in state.items():
            setattr(self, k, v)

    def __getstate__(self):
        return self.__dict__


def _load_generic_model(model_path: str):
    # Register stub so pickle can reconstruct CellTypist Model objects
    try:
        import celltypist.models as _ct_models
        ct_model_class = _ct_models.Model
    except ImportError:
        ct_model_class = _CellTypistModelStub

    # Try joblib first
    try:
        import joblib
        with warnings.catch_warnings():
            try:
                from sklearn.exceptions import InconsistentVersionWarning
                warnings.simplefilter("ignore", InconsistentVersionWarning)
            except ImportError:
                pass
            return joblib.load(model_path)
    except Exception:
        pass

    # Try pickle with custom unpickler for CellTypist Model class
    class _CellTypistUnpickler(pickle.Unpickler):
        def find_class(self, module: str, name: str):
            if "Model" in name and "celltypist" in module:
                return ct_model_class
            return super().find_class(module, name)

    with open(model_path, "rb") as handle:
        try:
            return _CellTypistUnpickler(handle).load()
        except Exception:
            handle.seek(0)
            return pickle.load(handle)


def _extract_model_parts(model_obj) -> Tuple[object, object, List[str], List[str]]:
    """Extract classifier, scaler, features, labels from various model formats."""
    def first_not_none(*values):
        # Do not use ``a or b`` here: CellTypist stores features/classes as
        # NumPy arrays, whose truth value is intentionally ambiguous.
        return next((value for value in values if value is not None), None)

    classifier = None
    scaler = None
    features = None
    labels = None

    # CellTypist format: dict with 'Model' key containing a Model object
    if isinstance(model_obj, dict):
        inner = first_not_none(model_obj.get("Model"), model_obj.get("model"))
        if inner is not None:
            classifier = first_not_none(
                getattr(inner, "classifier", None),
                getattr(inner, "clf", None),
                inner if hasattr(inner, "predict_proba") or hasattr(inner, "classes_") else None,
            )
            scaler = getattr(inner, "scaler", None)
            features = first_not_none(getattr(inner, "features", None), getattr(inner, "genes", None))
            labels = first_not_none(getattr(inner, "cell_types", None), getattr(inner, "labels", None))
        # Fill each missing part from direct dictionary keys. Standard
        # CellTypist .pkl files use Model + Scaler_.
        if classifier is None:
            classifier = first_not_none(model_obj.get("classifier"), model_obj.get("model"))
        if scaler is None:
            scaler = first_not_none(model_obj.get("scaler"), model_obj.get("Scaler_"))
        if features is None:
            features = first_not_none(model_obj.get("features"), model_obj.get("genes"))
        if labels is None:
            labels = first_not_none(model_obj.get("cell_types"), model_obj.get("labels"))

    # Object with attributes (CellTypist Model class or similar)
    if classifier is None:
        classifier = first_not_none(getattr(model_obj, "classifier", None), getattr(model_obj, "clf", None))
    if scaler is None:
        scaler = getattr(model_obj, "scaler", None)
    if features is None:
        features = first_not_none(getattr(model_obj, "features", None), getattr(model_obj, "genes", None))
    if labels is None:
        labels = first_not_none(getattr(model_obj, "cell_types", None), getattr(model_obj, "labels", None))

    # Try nested model attribute
    if classifier is None and hasattr(model_obj, "model"):
        inner = getattr(model_obj, "model")
        classifier = first_not_none(getattr(inner, "classifier", None), inner)
        if features is None:
            features = first_not_none(getattr(inner, "features", None), getattr(inner, "genes", None))

    # Last resort: if it has classes_, it might be the classifier itself
    if classifier is None and hasattr(model_obj, "classes_"):
        classifier = model_obj

    if labels is None and classifier is not None and hasattr(classifier, "classes_"):
        labels = list(classifier.classes_)

    if classifier is None or features is None:
        raise RuntimeError(
            f"Unsupported CellTypist model format. "
            f"Type: {type(model_obj).__name__}, "
            f"Keys: {list(model_obj.keys()) if isinstance(model_obj, dict) else dir(model_obj)[:20]}"
        )

    features = [str(x) for x in list(features)]
    labels = [str(x) for x in list(labels)] if labels is not None else [str(x) for x in list(classifier.classes_)]
    return classifier, scaler, features, labels


def run_celltypist(adata, model_path: str, min_prob: float, min_delta: float):
    import numpy as np

    model_path = normalize_windows_path(model_path)
    if not model_path:
        raise RuntimeError("CellTypist requires a local model file")
    if not os.path.exists(model_path):
        raise RuntimeError(f"CellTypist model not found: {model_path}")
    # ``normalize()`` has already converted a validated count source to the
    # official CellTypist input scale.  A processed h5ad source is deliberately
    # left untouched, so validate that it is actually log1p-normalized instead
    # of accepting any arbitrary non-negative matrix.
    validate_celltypist_expression(adata)

    model_obj = _load_generic_model(model_path)
    classifier, scaler, features, labels = _extract_model_parts(model_obj)
    if not hasattr(classifier, "coef_") or not hasattr(classifier, "intercept_"):
        raise RuntimeError("CellTypist classifier is missing coefficient data")

    # Match only genes present in both data and model. Official CellTypist uses
    # exact feature names, keeps input-gene order, and drops model genes absent
    # from the query. Filling missing genes or matching case-insensitively would
    # change the trained linear score (and can cross-match human/mouse symbols).
    model_indices, data_indices = _match_celltypist_features(adata.var_names, features)

    used_features = len(model_indices)
    if not used_features:
        raise RuntimeError("None of the model features overlap with the current dataset")

    coef = np.asarray(classifier.coef_, dtype=np.float32)
    intercept = np.asarray(classifier.intercept_, dtype=np.float32).reshape(-1)
    if coef.ndim != 2 or coef.shape[1] != len(features):
        raise RuntimeError("CellTypist model feature and coefficient dimensions do not match")
    matched_coef = coef[:, model_indices]

    means = np.zeros(used_features, dtype=np.float32)
    scales = np.ones(used_features, dtype=np.float32)
    if scaler is not None:
        if not hasattr(scaler, "scale_"):
            raise RuntimeError("CellTypist scaler is missing scale data")
        scaler_scales = np.asarray(scaler.scale_, dtype=np.float32)
        if scaler_scales.size != len(features):
            raise RuntimeError("CellTypist scaler and model feature dimensions do not match")
        scales = scaler_scales[model_indices]
        scales[scales == 0] = 1.0
        if getattr(scaler, "with_mean", True) and getattr(scaler, "mean_", None) is not None:
            means = np.asarray(scaler.mean_, dtype=np.float32)[model_indices]

    class_labels = [str(value) for value in getattr(classifier, "classes_", labels)]
    if coef.shape[0] not in (1, len(class_labels)):
        raise RuntimeError("CellTypist class labels and coefficient dimensions do not match")

    cell_label_ids = np.empty(adata.n_obs, dtype=np.uint32)
    best_scores = np.empty(adata.n_obs, dtype=np.float32)
    confidence_gaps = np.empty(adata.n_obs, dtype=np.float32)
    label_dictionary = list(class_labels)
    label_to_id = {label: index for index, label in enumerate(label_dictionary)}
    unknown_id = label_to_id.get("Unknown")
    chunk_size = 4096
    for start in range(0, adata.n_obs, chunk_size):
        stop = min(start + chunk_size, adata.n_obs)
        X = np.asarray(to_dense(adata[start:stop, data_indices].X), dtype=np.float32)
        X = (X - means) / scales
        # CellTypist clips only large positive standardized values.
        np.minimum(X, 10.0, out=X)
        decision = X @ matched_coef.T + intercept
        decision = np.asarray(decision, dtype=np.float32)
        if decision.ndim == 1:
            decision = decision[:, None]

        if decision.shape[1] == 1 and len(class_labels) == 2:
            positive = 1.0 / (1.0 + np.exp(-np.clip(decision[:, 0], -80, 80)))
            probs = np.column_stack((1.0 - positive, positive))
        else:
            # CellTypist reports independent sigmoid probabilities rather than
            # sklearn's multiclass softmax probabilities.
            probs = 1.0 / (1.0 + np.exp(-np.clip(decision, -80, 80)))

        best_idx = probs.argmax(axis=1)
        best_prob = probs[np.arange(probs.shape[0]), best_idx]
        if probs.shape[1] > 1:
            second_prob = np.partition(probs, -2, axis=1)[:, -2]
        else:
            second_prob = np.zeros(probs.shape[0], dtype=np.float32)
        delta = best_prob - second_prob
        for row in range(probs.shape[0]):
            label = class_labels[int(best_idx[row])]
            if float(best_prob[row]) < min_prob or float(delta[row]) < min_delta:
                label = "Unknown"
                if unknown_id is None:
                    unknown_id = len(label_dictionary)
                    label_dictionary.append("Unknown")
                    label_to_id["Unknown"] = unknown_id
            cell_label_ids[start + row] = label_to_id[label]
            best_scores[start + row] = best_prob[row]
            confidence_gaps[start + row] = delta[row]

    return {
        "cell_label_ids": cell_label_ids,
        "cell_scores_f32": best_scores,
        "cell_confidences_f32": confidence_gaps,
        "label_dictionary": label_dictionary,
        "evidence_count": used_features,
    }


def summarize_clusters_by_votes(
    cluster_ids: List[int],
    cluster_names: List[str],
    cell_labels,
    cell_scores,
    cell_conf,
    label_dictionary=None,
):
    """Summarize either legacy string labels or compact numeric label IDs."""
    cluster_count = len(cluster_names)
    vote_counts: List[Dict[str, int]] = [{} for _ in range(cluster_count)]
    score_sums: List[Dict[str, float]] = [{} for _ in range(cluster_count)]
    cell_counts = [0] * cluster_count
    for cluster_id, label_value, score in zip(cluster_ids, cell_labels, cell_scores):
        cluster_index = int(cluster_id)
        if cluster_index < 0 or cluster_index >= cluster_count:
            continue
        if label_dictionary is not None:
            label_index = int(label_value)
            label = label_dictionary[label_index] if 0 <= label_index < len(label_dictionary) else "Unknown"
        else:
            label = str(label_value)
        cell_counts[cluster_index] += 1
        vote_counts[cluster_index][label] = vote_counts[cluster_index].get(label, 0) + 1
        score_sums[cluster_index][label] = score_sums[cluster_index].get(label, 0.0) + float(score)

    summaries = []
    for cluster_index, cluster_name in enumerate(cluster_names):
        n_cells = cell_counts[cluster_index]
        if n_cells == 0:
            summaries.append({
                "cluster_index": cluster_index,
                "cluster_name": cluster_name,
                "label": "Unknown",
                "score": 0.0,
                "confidence": 0.0,
                "n_cells": 0,
            })
            continue
        cluster_votes = vote_counts[cluster_index]
        cluster_scores = score_sums[cluster_index]
        ordered = sorted(
            cluster_votes.items(),
            key=lambda item: (
                -item[1],
                -(cluster_scores.get(item[0], 0.0) / item[1]),
                item[0],
            ),
        )
        top_label, top_count = ordered[0]
        summaries.append({
            "cluster_index": cluster_index,
            "cluster_name": cluster_name,
            "label": top_label,
            "score": cluster_scores[top_label] / top_count,
            "confidence": top_count / n_cells,
            "n_cells": n_cells,
        })
    return summaries


def summarize_clusters_sctype(
    cluster_ids: List[int],
    cluster_names: List[str],
    cluster_score_sums,
    score_labels: List[str],
    cluster_cell_counts=None,
):
    """Cluster-level summarization matching original scType.

    - Sum ES scores across all cells in a cluster
    - Top-scoring cell type is the cluster label
    - If top summed score < ncells / 4, label as Unknown
    """
    import numpy as np

    cluster_count = len(cluster_names)
    # ``cluster_score_sums`` is produced incrementally by run_sctype.  Accept
    # the old cells-by-label matrix as a compatibility fallback for callers
    # outside this worker, but never create it in the normal path.
    if cluster_score_sums is None:
        return []
    score_array = np.asarray(cluster_score_sums)
    if cluster_cell_counts is None and score_array.ndim == 2 and score_array.shape[0] == len(cluster_ids):
        cluster_array = np.asarray(cluster_ids, dtype=np.intp)
        valid = (cluster_array >= 0) & (cluster_array < cluster_count)
        cell_counts = np.bincount(cluster_array[valid], minlength=cluster_count)
        summed_by_cluster = np.zeros((cluster_count, score_array.shape[1]), dtype=np.float64)
        np.add.at(summed_by_cluster, cluster_array[valid], score_array[valid])
    else:
        summed_by_cluster = score_array
        cell_counts = (
            np.asarray(cluster_cell_counts, dtype=np.intp)
            if cluster_cell_counts is not None
            else np.zeros(cluster_count, dtype=np.intp)
        )

    summaries = []
    for cluster_index, cluster_name in enumerate(cluster_names):
        n_cells = int(cell_counts[cluster_index])
        if n_cells == 0:
            summaries.append({
                "cluster_index": cluster_index,
                "cluster_name": cluster_name,
                "label": "Unknown",
                "score": 0.0,
                "confidence": 0.0,
                "n_cells": 0,
            })
            continue

        summed_scores = summed_by_cluster[cluster_index]
        top_idx = int(np.argmax(summed_scores))
        top_score = float(summed_scores[top_idx])
        second_score = float(np.partition(summed_scores, -2)[-2]) if len(score_labels) > 1 else 0.0
        mean_top_score = top_score / n_cells
        label = score_labels[top_idx]

        # Original scType threshold: summed score < ncells / 4
        if top_score < n_cells / 4.0:
            label = "Unknown"

        summaries.append({
            "cluster_index": cluster_index,
            "cluster_name": cluster_name,
            "label": label,
            "score": mean_top_score,
            "confidence": top_score - second_score,
            "n_cells": n_cells,
        })
    return summaries


def _compact_annotation_result(result, include_cells: bool):
    """Convert worker arrays to a small JSON-compatible result contract.

    New workers return numeric numpy arrays and a label dictionary.  The
    compatibility branch keeps accepting the pre-compact string-list result
    used by older resident workers and unit-test fixtures.
    """
    import numpy as np

    if "cell_label_ids" not in result:
        return {
            "cell_labels": result.get("cell_labels", []) if include_cells else [],
            "cell_scores": result.get("cell_scores", []) if include_cells else [],
            "cell_confidences": result.get("cell_confidences", []) if include_cells else [],
        }

    label_ids = np.asarray(result.get("cell_label_ids", []), dtype=np.uint32)
    label_dtype = "u16" if len(result.get("label_dictionary", [])) <= 65_535 else "u32"
    if label_dtype == "u16":
        label_ids = label_ids.astype(np.uint16, copy=False)
    scores = np.asarray(result.get("cell_scores_f32", []), dtype=np.float32)
    confidences = np.asarray(result.get("cell_confidences_f32", []), dtype=np.float32)
    return {
        "label_dictionary": [str(value) for value in result.get("label_dictionary", [])],
        "cell_label_ids": label_ids.tolist() if include_cells else [],
        "cell_label_ids_dtype": label_dtype,
        "cell_scores_f32": scores.tolist() if include_cells else [],
        "cell_confidences_f32": confidences.tolist() if include_cells else [],
    }


def _write_compact_result_frame(result_path: str, value: dict):
    """Write DNAR v1: JSON metadata + native little-endian numeric arrays."""
    import numpy as np

    label_dictionary = [str(label) for label in value.get("label_dictionary", [])]
    dtype = "u16" if len(label_dictionary) <= 65_535 else "u32"
    ids = np.asarray(value.get("cell_label_ids", []), dtype=np.uint16 if dtype == "u16" else np.uint32)
    scores = np.asarray(value.get("cell_scores_f32", []), dtype="<f4")
    confidences = np.asarray(value.get("cell_confidences_f32", []), dtype="<f4")
    if not (len(ids) == len(scores) == len(confidences)):
        raise RuntimeError("Compact annotation arrays have inconsistent lengths")

    descriptors = {}
    offset = 0
    for name, array, array_dtype in (
        ("cell_label_ids", ids, dtype),
        ("cell_scores_f32", scores, "f32"),
        ("cell_confidences_f32", confidences, "f32"),
    ):
        descriptors[name] = {"dtype": array_dtype, "offset": offset, "count": int(array.size)}
        offset += int(array.nbytes)

    metadata = dict(value)
    metadata.pop("cell_label_ids", None)
    metadata.pop("cell_scores_f32", None)
    metadata.pop("cell_confidences_f32", None)
    metadata["cell_label_ids_dtype"] = dtype
    metadata["cell_count"] = int(ids.size)
    metadata["arrays"] = descriptors
    if not (np.isfinite(scores).all() and np.isfinite(confidences).all()):
        raise RuntimeError("Compact annotation scores contain non-finite values")
    metadata_bytes = json.dumps(metadata, separators=(",", ":"), allow_nan=False).encode("utf-8")
    payload_length = offset
    frame_header = b"DNAR" + struct.pack("<III", 1, len(metadata_bytes), payload_length)
    with open(result_path, "wb") as handle:
        handle.write(frame_header)
        handle.write(metadata_bytes)
        for array in (ids, scores, confidences):
            handle.write(array.tobytes(order="C"))


def load_annotation_inputs(indices_path, config):
    """Load legacy JSON or the compact DNAN per-cell transport."""
    if indices_path.lower().endswith(".bin"):
        import struct
        import numpy as np

        with open(indices_path, "rb") as handle:
            header = handle.read(16)
            if len(header) != 16 or header[:4] != b"DNAN":
                raise RuntimeError("Invalid annotation binary frame")
            version, metadata_len, payload_len = struct.unpack("<III", header[4:])
            if version != 1 or metadata_len <= 0:
                raise RuntimeError("Unsupported annotation binary frame")
            metadata_bytes = handle.read(metadata_len)
            if len(metadata_bytes) != metadata_len:
                raise RuntimeError("Annotation binary metadata is truncated")
            metadata = json.loads(metadata_bytes.decode("utf-8"))
            payload_offset = 16 + metadata_len
        if os.path.getsize(indices_path) != payload_offset + payload_len:
            raise RuntimeError("Annotation binary payload is truncated")
        try:
            indices_meta = metadata["indices"]
            clusters_meta = metadata["clusters"]
            raw_cell_count = metadata["cellCount"]
            cell_count = int(raw_cell_count)
            if isinstance(raw_cell_count, bool) or raw_cell_count != cell_count or cell_count <= 0:
                raise ValueError
            if (
                indices_meta["dtype"] != "u32"
                or clusters_meta["dtype"] != "u8"
                or int(indices_meta["count"]) != cell_count
                or int(clusters_meta["count"]) != cell_count
            ):
                raise ValueError
            indices_offset = int(indices_meta["offset"])
            clusters_offset = int(clusters_meta["offset"])
            indices_bytes = cell_count * 4
            clusters_bytes = cell_count
            if (
                indices_offset < 0
                or clusters_offset < 0
                or indices_offset + indices_bytes > payload_len
                or clusters_offset + clusters_bytes > payload_len
                or indices_offset + indices_bytes > clusters_offset
            ):
                raise ValueError
        except (KeyError, TypeError, ValueError, OverflowError) as error:
            raise RuntimeError("Annotation binary frame has invalid array descriptors") from error
        cell_indices = np.memmap(
            indices_path,
            mode="r",
            dtype="<u4",
            offset=payload_offset + indices_offset,
            shape=(cell_count,),
        )
        config["cluster_ids"] = np.memmap(
            indices_path,
            mode="r",
            dtype="u1",
            offset=payload_offset + clusters_offset,
            shape=(cell_count,),
        )
    else:
        with open(indices_path, encoding="utf-8") as handle:
            cell_indices = json.load(handle)
    return cell_indices, config


def main():
    if len(sys.argv) != 5:
        raise RuntimeError("Usage: run_annotation.py <source> <format> <annotation_input> <config.json>")

    source, data_format, indices_path, config_path = sys.argv[1:]
    with open(config_path, encoding="utf-8") as handle:
        config = json.load(handle)
    cell_indices, config = load_annotation_inputs(indices_path, config)

    source_adata = load_source(source, data_format)
    try:
        adata, normalize_input = select_expression_adata(source_adata, cell_indices)
    finally:
        if getattr(source_adata, "isbacked", False):
            source_adata.file.close()
        del source_adata
    adata = normalize(adata, normalize_input)
    if adata.n_obs == 0:
        raise RuntimeError("Cell annotation requires at least one selected cell")

    method = str(config.get("method", "sctype")).lower()
    level = str(config.get("level", "both")).lower()
    tissue = str(config.get("tissue", "Immune system"))
    # Zero thresholds preserve CellTypist's official ``best match`` behavior.
    # DNBCScope can opt into Unknown labels through the confidence filter in
    # the annotation panel.
    min_delta = float(config.get("min_delta", 0.0))
    cluster_ids = config.get("cluster_ids", [])
    cluster_names = [str(v) for v in config.get("cluster_names", [])]
    has_cluster_ids = len(cluster_ids) > 0
    if level not in ("cell", "cluster", "both"):
        raise RuntimeError(f"Unsupported annotation level: {level}")
    if has_cluster_ids:
        if not cluster_names:
            raise RuntimeError("Cluster names are required when cluster IDs are provided")
        cluster_ids = _validated_cluster_ids(cluster_ids, adata.n_obs, len(cluster_names))

    # Always annotate individual cells first. Cluster results are summaries of
    # those scores/votes over the existing project clusters; predicting from an
    # average expression vector would not match either method's cell-level
    # semantics. This is a DNBCScope cluster consensus, not CellTypist's
    # optional Leiden over-clustering majority-voting pipeline.
    include_cells = level in ("cell", "both")
    if method == "sctype":
        result = run_sctype(
            adata,
            tissue,
            cluster_ids if has_cluster_ids else None,
            cluster_names if has_cluster_ids else None,
        )
    elif method == "celltypist":
        result = run_celltypist(
            adata,
            str(config.get("model_path", "")),
            float(config.get("min_probability", 0.0)),
            min_delta,
        )
    else:
        raise RuntimeError(f"Unsupported annotation method: {method}")

    cluster_summaries = []
    if level in ("cluster", "both") and has_cluster_ids and cluster_names:
        if method == "sctype":
            cluster_summaries = summarize_clusters_sctype(
                cluster_ids,
                cluster_names,
                result.get("cluster_score_sums"),
                result["score_labels"],
                result.get("cluster_cell_counts"),
            )
        else:
            labels = result.get("cell_labels", result.get("cell_label_ids", []))
            cluster_summaries = summarize_clusters_by_votes(
                cluster_ids,
                cluster_names,
                labels,
                result.get("cell_scores", result.get("cell_scores_f32", [])),
                result.get("cell_confidences", result.get("cell_confidences_f32", [])),
                result.get("label_dictionary"),
            )

    compact = _compact_annotation_result(result, include_cells)
    output = {
        "method": method,
        "result_level": level,
        "tissue": tissue if method == "sctype" else "",
        "evidence_count": result["evidence_count"],
        "cluster_summaries": cluster_summaries,
        "cluster_aggregation": "current-cluster-consensus" if cluster_summaries else "",
    }
    output.update(compact)
    emit_json_result(output)


if __name__ == "__main__":
    main()
